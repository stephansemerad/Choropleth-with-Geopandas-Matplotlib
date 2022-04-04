import os
import random
import pandas as pd
import seaborn as sns
import geopandas as gpd
import matplotlib.pyplot as plt
import matplotlib.colors as colors
import numpy as np
from openpyxl import load_workbook

class EasyChoroplethMap:
    def __init__(self):
        self.data_path      = None
        self.shape_path     = "./easy_choroplet_map/shapefiles/Europe.shp"
        self.map_df         = gpd.read_file(self.shape_path)
        self.color_map      = 'RdPu'
        
        self.title          = ''
        self.subtitle       = ''
        self.source         = ''
        
    def random_data(self):
        self.map_df['DATA'] = np.random.randint(1, 600, self.map_df.shape[0])

    def load_excel_data(self, file_path):
        if not os.path.exists(file_path): raise Exception(f'Could not find {file_path}, please check the filepath provided')

        wb = load_workbook(file_path, data_only=True)
        
        if 'data' not in wb.sheetnames: raise Exception(f'Could not sheet data in workbook')
        ws = wb['data']
        
        self.title              = ws['c2'].value
        self.subtitle           = ws['c3'].value
        self.source             = ws['c4'].value
        self.year_for_map       = ws['c5'].value
        self.year_for_column    = ws['c6'].value
        
        pass    
    
    def create_map(self, file_path):
        sns.set(style='whitegrid', palette='pastel', color_codes=True)

        fig, ax = plt.subplots(1, figsize=(15,10))
        ax.set_xticks([])
        ax.set_yticks([])

        cmap = plt.get_cmap('RdPu')
        new_cmap = self.truncate_colormap(cmap, 0.3, 1)
        
        map_df = self.map_df.set_index('NAME')
        map_df.plot(column='DATA', cmap=new_cmap, scheme='QUANTILES', k=6,
                    linewidth=0.9,
                    ax=ax,
                    edgecolor='1',
                    legend=True, 
                    legend_kwds={'loc': 'upper left'}
                    )

        map_df['coords'] = map_df['geometry'].apply(lambda x: x.representative_point().coords[:])
        map_df['coords'] = [coords[0] for coords in map_df['coords']]
        for index, row in map_df.iterrows():
            plt.annotate(text=index, xy=row['coords'], horizontalalignment='center', size=5, color='black')

        plt.title(f'{self.title}\n{self.subtitle}', loc='left', size="10")
        plt.xlabel(f'source:  {self.source}', loc="left", size="8")
        plt.savefig(file_path)
        print(f'created file {file_path}')

    def unique_list(self):
        unique_list = self.map_df.index.unique()
        unique_list = [x for x in unique_list]

    def truncate_colormap(self, cmap, minval=0.0, maxval=1.0, n=100):
        new_cmap = colors.LinearSegmentedColormap.from_list('trunc({n},{a:.2f},{b:.2f})'.format(n=cmap.name, a=minval, b=maxval),cmap(np.linspace(minval, maxval, n)))
        return new_cmap
